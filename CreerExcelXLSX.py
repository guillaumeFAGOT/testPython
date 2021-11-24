#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Oct 22 09:12:16 2021

@author: guillaume
""" 
# Chargement des librairies de gestion Python
from openpyxl import load_workbook
from openpyxl.chart import (
    AreaChart,
    Reference,
    Series,
)
from openpyxl import Workbook

# fonction de création d'un classeur et de sauvegarde sur le repertoire courant
def CreerClasseur():
    wb = Workbook()
    ws = wb.active
    ws1 = wb.create_sheet("Mysheet")
    wb.save('balances.xlsx')
    
#Aller chercher les valeurs de la feuille1 "Sheet"    
def AllerChercherDonneesClasseur():

    wb = load_workbook('balances.xlsx')
    ws = wb["Sheet"]
    # selection dans la feuille courante de la plage ligne 1 à ligne 6 jusqu a la colonne 6
    for row in ws.iter_rows(min_row=1, max_col=6, max_row=6):
        for cell in row:
            if cell.value is not None:
                print (cell.value)
            
                
#AllerChercherDonneesClasseur()
#Ajout d un graphe dans le classeur
def Ajout_D_Un_Graphe():

    
    wb = Workbook()
    ws = wb.active
    
    # integration des donnees
    rows = [
    ['Number', 'Batch 1', 'Batch 2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10],
]
    
    for row in rows:
        ws.append(row)

    # Ajout d un graphique    
    chart = AreaChart()
    chart.title = "Area Chart"
    chart.style = 13
    chart.x_axis.title = 'Test'
    chart.y_axis.title = 'Percentage'
    
    # la fonction Reference permet de selectionner une plage : les titres et les donnees
    cats = Reference(ws, min_col=1, min_row=1, max_row=7)
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "A10")
    
    wb.save("area.xlsx")
 
# Appel de la fonction
Ajout_D_Un_Graphe()
    
    
    
